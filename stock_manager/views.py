import logging
from django.conf import settings
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from rest_framework import viewsets
from rest_framework.permissions import IsAuthenticated
from .models import Item, ShopItem, TransferItem, Admin
from .serializers import ItemSerializer, ShopItemSerializer, TransferItemSerializer
from .pagination import CustomPagination
from django.contrib.auth.models import User  # For accessing the User model
from rest_framework.response import (
    Response,
)  # For returning HTTP responses in REST framework
from rest_framework.decorators import api_view
from django.db.models.functions import Lower, Cast
from rest_framework.decorators import permission_classes
from django.http import JsonResponse, FileResponse
from rest_framework.permissions import IsAuthenticated
from rest_framework import status
from rest_framework.exceptions import ValidationError
from rest_framework.decorators import action
from django.views.decorators.csrf import csrf_exempt, ensure_csrf_cookie
from django.db.models import IntegerField, Q, ForeignKey, OneToOneField, ManyToManyField
from email_service.email import SendEmail
from io import BytesIO
from openpyxl import Workbook, load_workbook
from functools import reduce
from django.db import transaction

logger = logging.getLogger(__name__)


# API View
class ItemViewSet(viewsets.ModelViewSet):
    queryset = Item.objects.all()
    serializer_class = ItemSerializer
    lookup_field = "sku"
    permission_classes = [IsAuthenticated]
    pagination_class = CustomPagination

    def get_queryset(self):
        queryset = Item.objects.all()
        search_query = self.request.query_params.get("search", None)
        if search_query:
            queryset = queryset.filter(
                Q(description__icontains=search_query) | Q(sku__icontains=search_query)
            )  # 🔍 Search filter
        ordering = self.request.query_params.get("ordering", None)
        if ordering:
            if ordering.startswith("-"):
                field = ordering[1:]
                if field == "quantity":
                    queryset = queryset.order_by(Cast(field, IntegerField()).desc())
                else:
                    queryset = queryset.order_by(Lower(field)).reverse()
            else:
                if ordering == "quantity":
                    queryset = queryset.order_by(Cast(ordering, IntegerField()))
                else:
                    queryset = queryset.order_by(Lower(ordering))
        else:
            queryset = queryset.order_by("last_updated").reverse()
        return queryset

    def update(self, request, *args, **kwargs):
        if not request.user.groups.filter(name="managers").exists():
            return Response(
                {"detail": "Permission denied."}, status=status.HTTP_403_FORBIDDEN
            )
        try:
            item = Item.objects.get(sku=request.data.get("sku"))
        except Item.DoesNotExist:
            return Response(
                {"error": "Item not found."}, status=status.HTTP_404_NOT_FOUND
            )
        serializer = ItemSerializer(item, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()  # Save the changes to the database
            return Response(serializer.data, status=status.HTTP_200_OK)
        else:
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def destroy(self, request, *args, **kwargs):
        if not request.user.groups.filter(name="managers").exists():
            return Response(
                {"detail": "Permission denied."}, status=status.HTTP_403_FORBIDDEN
            )
        return super().destroy(request, *args, **kwargs)


class ShopItemViewSet(viewsets.ModelViewSet):
    queryset = ShopItem.objects.all()
    serializer_class = ShopItemSerializer
    lookup_field = "item__sku"
    permission_classes = [IsAuthenticated]
    pagination_class = CustomPagination

    def get_queryset(self):
        queryset = ShopItem.objects.filter(shop_user=self.request.user)
        search_query = self.request.query_params.get("search", None)
        if search_query:
            queryset = queryset.filter(
                Q(item__description__icontains=search_query)
                | Q(item__sku__icontains=search_query)
            )  # 🔍 Search filter
        ordering = self.request.query_params.get("ordering", None)
        if ordering.startswith("-"):
            field = ordering[1:]
            if field == "quantity":
                queryset = queryset.order_by(Cast(field, IntegerField()).desc())
            else:
                queryset = queryset.order_by(Lower(field)).reverse()
        else:
            if ordering == "quantity":
                queryset = queryset.order_by(Cast(ordering, IntegerField()))
            else:
                queryset = queryset.order_by(Lower(ordering))
        return queryset


class TransferItemViewSet(viewsets.ModelViewSet):
    queryset = TransferItem.objects.all()
    serializer_class = TransferItemSerializer
    lookup_field = "item__sku"
    permission_classes = [IsAuthenticated]
    pagination_class = CustomPagination

    def get_queryset(self):
        user = self.request.user
        if user.groups.filter(name="managers").exists():
            queryset = TransferItem.objects.filter(ordered=True)
        else:
            queryset = TransferItem.objects.filter(shop_user=user)
        search_query = self.request.query_params.get("search", None)
        if search_query:
            queryset = queryset.filter(
                Q(item__description__icontains=search_query)
                | Q(item__sku__icontains=search_query)
            )  # 🔍 Search filter
        ordering = self.request.query_params.get("ordering", None)
        if ordering.startswith("-"):
            field = ordering[1:]
            if field == "quantity":
                queryset = queryset.order_by(Cast(field, IntegerField()).desc())
            else:
                queryset = queryset.order_by(Lower(field)).reverse()
        else:
            if ordering == "quantity":
                queryset = queryset.order_by(Cast(ordering, IntegerField()))
            else:
                queryset = queryset.order_by(Lower(ordering))
        return queryset


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def set_edit_lock_status(request):
    if not request.user.groups.filter(name="managers").exists():
        return Response(
            {"detail": "Permission denied."}, status=status.HTTP_403_FORBIDDEN
        )

    edit_lock_status = request.data.get("edit_lock_status", False)
    admin, created = Admin.objects.get_or_create(id=1)
    admin.edit_lock = edit_lock_status
    admin.save()
    return Response(
        {"edit_lock": admin.edit_lock},
        status=status.HTTP_200_OK,
    )


@csrf_exempt
def get_edit_lock_status(request):
    if request.method == "GET":
        edit_lock = Admin.is_edit_locked()
        return JsonResponse({"edit_lock": edit_lock})
    return JsonResponse({"error": "Invalid request method"}, status=400)


# Main Page View
@ensure_csrf_cookie
@login_required
def index(request):
    return render(request, "index.html")


@api_view(["GET"])
def get_user(request):
    # Ensure the user is authenticated
    if not request.user.is_authenticated:
        return Response(
            {"detail": "Authentication credentials were not provided."}, status=401
        )

    user = request.user
    return Response(
        {
            "id": user.id,
            "username": user.username,
            "email": user.email,
            "groups": list(user.groups.values_list("name", flat=True)),
        }
    )


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def transfer_item(request):
    if Admin.objects.first().edit_lock:
        logger.debug("Transfer attempt while update mode is enabled.")
        return Response(
            {
                "detail": "Transfers are disabled as the warehouse is being maintained. Please try again later."
            },
            status=status.HTTP_403_FORBIDDEN,
        )
    if not request.user.groups.filter(name="shop_users").exists():
        logger.debug("Permission denied: user is not in shop_users group.")
        return Response(
            {"detail": "Permission denied."}, status=status.HTTP_403_FORBIDDEN
        )
    sku = request.data.get("sku")
    transfer_quantity = request.data.get("transfer_quantity")
    if not transfer_quantity.isdigit():
        logger.debug("Invalid transfer quantity: not an integer.")
        return Response(
            {"detail": "Transfer quantity must be an integer."},
            status=status.HTTP_400_BAD_REQUEST,
        )
    transfer_quantity = int(transfer_quantity)
    if transfer_quantity <= 0:
        logger.debug("Invalid transfer quantity: less than or equal to zero.")
        return Response(
            {"detail": "Transfer quantity must be greater than zero."},
            status=status.HTTP_400_BAD_REQUEST,
        )
    try:
        item = Item.objects.get(sku=sku)
        transfer_to_shop(item, request.user, transfer_quantity)
    except Item.DoesNotExist:
        logger.debug("Item not found: sku=%s", sku)
        return Response({"detail": "Item not found."}, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        logger.debug("Error during transfer: %s", str(e))
        return Response({"detail": str(e)}, status=status.HTTP_400_BAD_REQUEST)
    return Response({"detail": "Transfer successful."}, status=status.HTTP_200_OK)


def transfer_to_shop(
    item, shop_user, transfer_quantity, complete=False, cancel=False, manager=False
):
    if Admin.is_edit_locked() and not manager:
        raise ValueError(
            "Transfers are disabled as the warehouse is being maintained. Please try again later."
        )
    if cancel:
        transfer_item = TransferItem.objects.get(
            item=item, shop_user=shop_user
        ).delete()
    else:
        transfer_quantity = int(transfer_quantity)
        if item.quantity < transfer_quantity:
            raise ValueError("Not enough stock to transfer")
        if not complete:
            xfer_item, created = TransferItem.objects.get_or_create(
                shop_user=shop_user, item=item
            )
            if xfer_item.ordered:
                raise LookupError(
                    "This item has already been ordered and is awaiting dispatch. Please contact the warehouse manager if you wish to amend your order."
                )
            xfer_item.quantity = transfer_quantity
            xfer_item.save()
        else:
            if item.quantity < int(transfer_quantity):
                raise ValueError("Not enough stock to transfer")
            # transfer to ShopItem database
            shop_user = User.objects.get(id=shop_user)
            shop_item, created = ShopItem.objects.get_or_create(
                item=item, shop_user=shop_user
            )
            shop_item.quantity += transfer_quantity
            shop_item.save()
            # change quantity recorded for stock Item in warehouse
            item.quantity -= transfer_quantity
            item.save()
            # delete item from pending transfer
            TransferItem.objects.get(item=item, shop_user=shop_user).delete()


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def submit_transfer_request(request):
    try:
        if not Admin.is_edit_locked():
            queryset = TransferItem.objects.filter(
                shop_user=request.user.id, ordered=False
            )
            if queryset.exists():
                # send notification email
                SendEmail().compose(
                    records=list(
                        queryset.values(
                            "id",
                            "item__sku",
                            "item__description",
                            "item__retail_price",
                            "quantity",
                        )
                    ),
                    user=request.user,
                    notification_type=SendEmail.EmailType.STOCK_TRANSFER,
                )
                # update records ordered status to True
                queryset.update(ordered=True)
            else:
                return Response(
                    {"detail": "There were no outstanding items to request!"},
                    status=status.HTTP_400_BAD_REQUEST,
                )
        else:
            return Response(
                {
                    "detail": "Transfers cannot proceed while the warehouse is under maintenance. Please try again later."
                },
                status=status.HTTP_403_FORBIDDEN,
            )
    except Exception as e:
        logger.debug("Error while submitting transfer: %s", str(e))
        return Response({"detail": str(e)}, status=status.HTTP_400_BAD_REQUEST)
    return Response(
        {"detail": "Transfer successfully submitted."}, status=status.HTTP_200_OK
    )


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def complete_transfer(request):
    sku = request.data.get("sku")
    quantity = request.data.get("quantity")
    shop_user_id = request.data.get("shop_user_id")
    cancel = True if request.data.get("cancel") == "true" else False
    try:
        shop_user = User.objects.get(username=shop_user_id)
        shop_user_id = shop_user.id
    except User.DoesNotExist:
        logger.debug("User does not exist!")
        return Response(
            {"detail": "Shop user not found."}, status=status.HTTP_400_BAD_REQUEST
        )
    if not request.user.groups.filter(name="managers").exists() and not cancel:
        return Response(
            {"detail": "Permission denied. User is not in managers group."},
            status=status.HTTP_403_FORBIDDEN,
        )
    try:
        item = Item.objects.get(sku=sku)
        transfer_to_shop(
            manager=request.user.groups.filter(name="managers").exists(),
            item=item,
            shop_user=shop_user_id,
            transfer_quantity=quantity,
            complete=True,
            cancel=cancel,
        )
    except Item.DoesNotExist:
        logger.debug("Item not found: sku=%s", sku)
        return Response({"detail": "Item not found."}, status=status.HTTP_404_NOT_FOUND)
    except ValueError as e:
        logger.debug("ValueError during transfer: %s", str(e))
        return Response({"detail": str(e)}, status=status.HTTP_400_BAD_REQUEST)
    except Exception as e:
        return Response({"detail": str(e)}, status=status.HTTP_400_BAD_REQUEST)
    return Response(
        {"detail": "Transfer action successful."}, status=status.HTTP_200_OK
    )


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def export_data_excel(request):

    @staticmethod
    def get_related_field(obj, field_name):
        try:
            return reduce(
                lambda o, attr: getattr(o, attr, None) if o else None,
                field_name.split("__"),
                obj,
            )
        except AttributeError:
            return ""

    manager = request.user.groups.filter(name="managers").exists()
    shop_user = request.user.groups.filter(name="shop_users").exists()

    if not manager and not shop_user:
        logger.debug("Permission denied: user is not in shop_users or managers group.")
        return Response(
            {"detail": "Permission denied."}, status=status.HTTP_403_FORBIDDEN
        )

    workbook = Workbook()
    item_sheet = workbook.active
    item_sheet.title = "Warehouse Stock"
    item_fields = ["sku", "description", "retail_price", "quantity"]
    item_header = ["SKU", "Description", "Retail Price", "Quantity"]
    item_sheet.append(item_header)
    for item in Item.objects.only(*item_fields):
        row_data = [getattr(item, field, "") for field in item_fields]
        item_sheet.append(row_data)
    shop_item_sheet = workbook.create_sheet(title="Shop Stock")
    shop_item_relation_fields = ["shop_user", "item"]
    shop_item_retrieved_fields = [
        "shop_user__username",
        "item__sku",
        "item__description",
        "item__retail_price",
        "quantity",
    ]
    shop_item_header = [
        "Shop User",
        "SKU",
        "Description",
        "Retail Price",
        "Quantity",
    ]
    shop_item_queryset = ShopItem.objects.select_related(
        *shop_item_relation_fields
    ).only(*shop_item_retrieved_fields)
    shop_item_queryset = (
        shop_item_queryset.filter(**{"shop_user__username": request.user.username})
        if not manager
        else shop_item_queryset
    )
    shop_item_sheet.append(shop_item_header)
    for shop_item in shop_item_queryset:
        row_data = [
            get_related_field(shop_item, field) for field in shop_item_retrieved_fields
        ]
        shop_item_sheet.append(row_data)
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return FileResponse(
        output,
        as_attachment=True,
        filename="ssm_data.xlsx",
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def import_data_excel(request):
    if not request.user.groups.filter(name="managers").exists():
        logger.debug("Permission denied: user is not in managers group.")
        return Response({"detail": "Permission denied."}, status=403)
    
    if not getattr(settings, "ALLOW_UPLOADS", False):
        logger.debug("Attempted upload when disabled.")
        return Response({"detail": "Uploads are disabled in the app configuration."}, status=400)

    @staticmethod
    def field_changed(instance, field_name, new_value):
        """
        Check whether a field on an instance would change if updated with new_value.
        Uses the model field's to_python() to convert the new value.
        """
        try:
            field_obj = instance._meta.get_field(field_name)
            norm_new_value = field_obj.to_python(new_value)
        except Exception:
            norm_new_value = (
                new_value  # fallback if field not found or conversion fails
            )

        # Handle None vs. empty string
        old_value = getattr(instance, field_name)
        if old_value is None and norm_new_value in [None, ""]:
            return False
        return old_value != norm_new_value

    file = request.FILES.get("file")
    if not file or not file.name.endswith(".xlsx"):
        return Response(
            {"detail": "Invalid file format. Please upload an .xlsx file."},
            status=400,
        )

    # Define field mappings for Item and ShopItem models.
    item_field_mapping = {
        "SKU": "sku",
        "Description": "description",
        "Retail Price": "retail_price",
        "Quantity": "quantity",
    }
    shop_item_field_mapping = {
        "Shop User": "shop_user__username",
        "SKU": "item__sku",
        "Description": "item__description",
        "Retail Price": "item__retail_price",
        "Quantity": "quantity",
    }

    # We'll track which records are present in the Excel data.
    excel_item_skus = set()  # For Item (sku is the primary key)
    excel_shopitem_keys = (
        set()
    )  # For ShopItem (tuple of (shop_user.username, item.sku))

    try:
        workbook = load_workbook(file)
        with transaction.atomic():
            # Process "Warehouse Stock" sheet for Item model
            if "Warehouse Stock" in workbook.sheetnames:
                item_sheet = workbook["Warehouse Stock"]
                headers = [cell.value for cell in next(item_sheet.iter_rows(max_row=1))]
                for row in item_sheet.iter_rows(min_row=2, values_only=True):
                    # Build a data dictionary using the mapping
                    data = {
                        item_field_mapping[headers[i]]: value
                        for i, value in enumerate(row)
                        if headers[i] in item_field_mapping
                    }
                    sku = data.get("sku")
                    if not sku:
                        continue  # skip rows without an SKU
                    excel_item_skus.add(sku)
                    obj, created = Item.objects.get_or_create(sku=sku, defaults=data)
                    if not created:
                        updated = False
                        for key, value in data.items():
                            if field_changed(obj, key, value):
                                setattr(obj, key, value)
                                updated = True
                        if updated:
                            obj.save()

            # Process "Shop Stock" sheet for ShopItem model
            if "Shop Stock" in workbook.sheetnames:
                shop_item_sheet = workbook["Shop Stock"]
                headers = [
                    cell.value for cell in next(shop_item_sheet.iter_rows(max_row=1))
                ]
                for row in shop_item_sheet.iter_rows(min_row=2, values_only=True):
                    raw_data = {
                        shop_item_field_mapping[headers[i]]: value
                        for i, value in enumerate(row)
                        if headers[i] in shop_item_field_mapping
                    }
                    # Resolve related fields: fetch shop_user and item.
                    shop_username = raw_data.pop("shop_user__username", None)
                    item_sku = raw_data.pop("item__sku", None)
                    if not shop_username or not item_sku:
                        continue  # skip row if key data is missing

                    # Track the combination key
                    excel_shopitem_keys.add((shop_username, item_sku))

                    shop_user = User.objects.get(username=shop_username)
                    item = Item.objects.get(sku=item_sku)
                    obj, created = ShopItem.objects.get_or_create(
                        shop_user=shop_user, item=item
                    )

                    item_updated = False
                    shop_item_updated = False

                    for key, value in raw_data.items():
                        if key.startswith("item__"):
                            field = key.split("__", 1)[1]
                            if field_changed(item, field, value):
                                setattr(item, field, value)
                                item_updated = True
                        else:
                            if field_changed(obj, key, value):
                                setattr(obj, key, value)
                                shop_item_updated = True

                    if item_updated:
                        item.save()
                    if shop_item_updated:
                        obj.save()

            allow_delete = getattr(settings, "ALLOW_RECORD_DELETE_FROM_XLSX", False)
            if allow_delete:
                # Delete Items that are in the DB but not in the Excel file.
                # (Excel file is considered the source of truth.)
                deleted_items_count, _ = Item.objects.exclude(
                    sku__in=excel_item_skus
                ).delete()
                logger.debug(
                    "Deleted %s Item records not present in Excel", deleted_items_count
                )

                # Delete ShopItem records that are in the DB but not in the Excel file.
                # Iterate over all ShopItem records and delete those that do not match any Excel key.
                for shop_item in ShopItem.objects.select_related("shop_user", "item"):
                    key = (shop_item.shop_user.username, shop_item.item.sku)
                    if key not in excel_shopitem_keys:
                        shop_item.delete()

    except Exception as e:
        logger.debug("Error while importing Excel file: %s", str(e))
        return Response({"detail": str(e)}, status=400)

    return Response({"detail": "Data successfully imported."}, status=200)
