import logging
from functools import reduce
from io import BytesIO
from openpyxl import Workbook, load_workbook
from rest_framework.response import Response
from rest_framework import status
from django.http import FileResponse
from django.db import transaction
from datetime import datetime
import pytz

from .models import Item, ShopItem, User, Admin

logger = logging.getLogger(__name__)

try:
    from .custom_funcs import spreadsheet_convert
    HAS_SPREADSHEET_CONVERT = True
except ImportError:
    spreadsheet_convert = None
    HAS_SPREADSHEET_CONVERT = False

class SpreadsheetTools:
    def field_changed(self, instance, field_name, new_value):
        """
        Check whether a field on an instance would change if updated with new_value.
        Uses the model field's to_python() to convert the new value.
        """
        try:
            field_obj = instance._meta.get_field(field_name)
            norm_new_value = field_obj.to_python(new_value)
        except Exception:
            norm_new_value = new_value  # fallback if field not found or conversion fails

        old_value = getattr(instance, field_name)
        if old_value is None and norm_new_value in [None, ""]:
            return False
        return old_value != norm_new_value
    def __init__(self, request=None):
        self.request = request
        self.user = request.user

    def get_related_field(self, obj, field_name):
        try:
            return reduce(
                lambda o, attr: getattr(o, attr, None) if o else None,
                field_name.split("__"),
                obj,
            )
        except AttributeError:
            return ""

    def convert_custom_incoming_format(self, workbook):
        logger.info("convert_custom_incoming_format called. HAS_SPREADSHEET_CONVERT=%s", HAS_SPREADSHEET_CONVERT)
        if HAS_SPREADSHEET_CONVERT:
            try:
                logger.info("Calling convert_excel...")
                return spreadsheet_convert.convert_excel(workbook)
            except ValueError as ve:
                logger.error(f"Spreadsheet validation error: {ve}", exc_info=False)
                raise ve
            except Exception as e:
                logger.error(f"Error in convert_excel: {e}", exc_info=False)
                raise Exception(f"Custom spreadsheet conversion failed: {e}")
        else:
            raise Exception("A 'custom_funcs/spreadsheet_convert.py' file does not exist or could not be imported.")

    def handle_excel_upload(self):
        logger.info("handle_excel_upload called for user: %s", getattr(self.user, "username", "unknown"))
        self.cleanup_orphaned_shopitems()
        file_obj = self.request.FILES.get("file")
        if not file_obj or not file_obj.name.endswith(".xlsx"):
            return Response({"detail": "Invalid file format. Please upload an .xlsx file."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            in_memory_file = BytesIO(file_obj.read())
            in_memory_file.seek(0)
            workbook = load_workbook(in_memory_file)
            logger.info("Workbook loaded successfully.")
            nan_details = []
            for sheet_name in ["Warehouse Stock", "Shop Stock"]:
                if sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    headers = [cell.value for cell in next(sheet.iter_rows(max_row=1))]
                    nan_details.extend(self.check_nan_with_original(sheet, headers, logger))
            if nan_details:
                for row_num, sheet_name, val_type, val, sku_val in nan_details:
                    logger.error(f"NaN detected in Quantity field: Row {row_num} (SKU: {sku_val}) in sheet '{sheet_name}' (type: {val_type}) value={val!r}")
                return Response({"detail": "Invalid NaN values detected. Please ensure all quantity fields are valid numbers."}, status=400)
        except Exception as e:
            logger.error("Error while importing Excel file: %s", str(e), exc_info=False)
            return Response({"detail": "Failed to upload stock data."}, status=400)

        in_memory_file.seek(0)
        try:
            workbook = load_workbook(in_memory_file)
            with transaction.atomic():
                # --- Warehouse Stock ---
                item_field_mapping = {
                    "SKU": "sku",
                    "Description": "description",
                    "Retail Price": "retail_price",
                    "Quantity": "quantity",
                }
                excel_item_skus = set()
                if "Warehouse Stock" not in workbook.sheetnames:
                    try:
                        converted = self.convert_custom_incoming_format(workbook)
                        if "Warehouse Stock" in converted.sheetnames:
                            workbook = converted
                    except ValueError as ve:
                        logger.error(f"Spreadsheet validation error: {ve}", exc_info=False)
                        return Response({"detail": str(ve)}, status=400)
                    except Exception as e:
                        logger.warning(f"The Warehouse Stock sheet was missing. Conversion failed: {e}", exc_info=False)
                        return Response({"detail": str(e)}, status=400)
                if "Warehouse Stock" in workbook.sheetnames:
                    item_sheet = workbook["Warehouse Stock"]
                    headers = [cell.value for cell in next(item_sheet.iter_rows(max_row=1))]
                    if not all(col in headers for col in ["SKU", "Description", "Retail Price", "Quantity"]):
                        logger.warning("Default headers could not be mapped. Consulting custom mappings...")
                        item_sheet = self.convert_custom_incoming_format(workbook)["Warehouse Stock"]
                        headers = [cell.value for cell in next(item_sheet.iter_rows(max_row=1))]
                    for row in item_sheet.iter_rows(min_row=2, values_only=True):
                        data = {item_field_mapping[headers[i]]: value for i, value in enumerate(row) if headers[i] in item_field_mapping}
                        sku = data.get("sku")
                        if not sku:
                            continue
                        excel_item_skus.add(sku)
                        obj, created = Item.objects.get_or_create(sku=sku, defaults=data)
                        if not created:
                            updated = False
                            for key, value in data.items():
                                if self.field_changed(obj, key, value):
                                    setattr(obj, key, value)
                                    updated = True
                            if obj.is_active is False:
                                obj.is_active = True
                                updated = True
                            if updated:
                                obj.save()
                # --- Deactivate warehouse items not present in the spreadsheet if deletions allowed ---
                if Admin.is_allow_upload_deletions():
                    Item.objects.filter(is_active=True).exclude(sku__in=excel_item_skus).update(is_active=False)

                # --- Shop Stock ---
                shop_item_field_mapping = {
                    "Shop User": "shop_user__username",
                    "SKU": "item__sku",
                    "Description": "item__description",
                    "Retail Price": "item__retail_price",
                    "Quantity": "quantity",
                }
                unique_shop_items_in_excel = set()
                unique_shop_users_in_excel = set()
                if "Shop Stock" not in workbook.sheetnames:
                    try:
                        converted = self.convert_custom_incoming_format(workbook)
                        if "Shop Stock" in converted.sheetnames:
                            workbook = converted
                    except ValueError as ve:
                        logger.error(f"Spreadsheet validation error: {ve}", exc_info=False)
                        return Response({"detail": str(ve)}, status=400)
                    except Exception as e:
                        logger.warning(f"The Shop Stock sheet was missing. Conversion failed: {e}", exc_info=False)
                        return Response({"detail": str(e)}, status=400)
                if "Shop Stock" in workbook.sheetnames:
                    shop_item_sheet = workbook["Shop Stock"]
                    headers = [cell.value for cell in next(shop_item_sheet.iter_rows(max_row=1))]
                    if not all(col in headers for col in ["SKU", "Description", "Retail Price", "Quantity", "Shop User"]):
                        logger.warning("Default headers could not be mapped. Consulting custom mappings...")
                        shop_item_sheet = self.convert_custom_incoming_format(workbook)["Shop Stock"]
                        headers = [cell.value for cell in next(shop_item_sheet.iter_rows(max_row=1))]
                    for row in shop_item_sheet.iter_rows(min_row=2, values_only=True):
                        raw_data = {shop_item_field_mapping[headers[i]]: value for i, value in enumerate(row) if headers[i] in shop_item_field_mapping}
                        shop_username = raw_data.pop("shop_user__username", None)
                        item_sku = raw_data.pop("item__sku", None)
                        if not shop_username or not item_sku:
                            continue
                        unique_shop_items_in_excel.add(item_sku)
                        try:
                            shop_user = User.objects.get(username=shop_username)
                        except User.DoesNotExist:
                            logger.warning(f"Shop user '{shop_username}' not found. Skipping row.")
                            continue
                        unique_shop_users_in_excel.add(shop_user)
                        try:
                            item = Item.objects.get(sku=item_sku)
                        except Item.DoesNotExist:
                            item_defaults = {
                                "description": raw_data.get("item__description", ""),
                                "retail_price": raw_data.get("item__retail_price", 0.00) or 0.00,
                                "quantity": raw_data.get("quantity", 0) or 0,
                                "is_active": False,
                            }
                            item = Item.objects.create(sku=item_sku, **item_defaults)
                            logger.warning(f"Item with SKU '{item_sku}' not found. Created with is_active=False and defaults.")
                        obj, created = ShopItem.objects.get_or_create(shop_user=shop_user, item=item)
                        item_updated = False
                        shop_item_updated = False
                        for key, value in raw_data.items():
                            if key.startswith("item__"):
                                field = key.split("__", 1)[1]
                                if self.field_changed(item, field, value):
                                    setattr(item, field, value)
                                    item_updated = True
                            else:
                                if self.field_changed(obj, key, value):
                                    setattr(obj, key, value)
                                    shop_item_updated = True
                        if item_updated:
                            item.save()
                        if shop_item_updated:
                            obj.save()
                    # --- Delete ShopItems for missing (shop_user, item) only if deletions allowed ---
                    if Admin.is_allow_upload_deletions():
                        excel_shopitem_keys = set()
                        for row in shop_item_sheet.iter_rows(min_row=2, values_only=True):
                            row_dict = {headers[i]: value for i, value in enumerate(row) if headers[i] in shop_item_field_mapping}
                            shop_username = row_dict.get("Shop User")
                            item_sku = row_dict.get("SKU")
                            if shop_username and item_sku:
                                excel_shopitem_keys.add((shop_username, item_sku))
                        for shop_item in ShopItem.objects.select_related("shop_user", "item").filter(item__isnull=False):
                            key = (shop_item.shop_user.username, shop_item.item.sku)
                            if key not in excel_shopitem_keys:
                                try:
                                    shop_item.delete()
                                except Exception as ex:
                                    logger.error(f"Failed to delete ShopItem for user '{shop_item.shop_user.username}' and item '{shop_item.item.sku}': {ex}", exc_info=False)
                                    raise
            logger.info("handle_excel_upload completed successfully.")
        except Exception as e:
            logger.error("Unexpected error while processing Excel file: %s", str(e), exc_info=False)
            return Response({"detail": "An unexpected error occurred."}, status=400)

        return Response({"detail": "Data has been processed according to configuration."}, status=200)

    def cleanup_orphaned_shopitems(self):
        ShopItem.objects.filter(item__isnull=True).delete()
        orphaned_shopitems = ShopItem.objects.exclude(item__isnull=True).exclude(
            item_id__in=Item.objects.values_list("sku", flat=True)
        )
        count_orphans = orphaned_shopitems.count()
        orphaned_shopitems.delete()
        if count_orphans:
            logger.warning("Deleted %d orphaned ShopItem rows with invalid item_id", count_orphans)

    def check_nan_with_original(self, sheet, headers, logger):
        import pandas as pd
        nan_details = []
        if "Quantity" in headers:
            quantity_idx = headers.index("Quantity")
            sku_idx = headers.index("SKU") if "SKU" in headers else None
            rows = list(sheet.iter_rows(min_row=2, values_only=True))
            cleaned = [0 if pd.isna(val) or (isinstance(val, str) and val.strip() == "") else val for val in (row[quantity_idx] for row in rows)]
            cleaned_numeric = pd.to_numeric(cleaned, errors="coerce")
            for row_num, (val, row) in enumerate(zip(cleaned_numeric, rows), start=2):
                if pd.isna(val):
                    sku_val = row[sku_idx] if sku_idx is not None else None
                    logger.error(f"NaN TRIGGER: Row {row_num}, SKU={sku_val}, col=Quantity, value={val!r} (type={type(val).__name__})")
                    nan_details.append((row_num, sheet.title, type(val).__name__, val, sku_val))
        return nan_details
