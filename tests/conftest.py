"""
Shared fixtures for the SSM test suite.

Group / user fixtures mirror the roles the app expects to be created via
the admin site: 'managers', 'shop_users' and 'receive_mail'.
"""

from io import BytesIO

import pytest
from django.contrib.auth.models import Group, User
from django.core.files.uploadedfile import SimpleUploadedFile
from openpyxl import Workbook, load_workbook
from rest_framework.test import APIClient

from stock_manager.models import Admin, Item

# Default sheet headers as produced by the app's own Excel export.
WAREHOUSE_HEADER = ["SKU", "Description", "Retail Price", "Quantity"]
SHOP_HEADER = ["Shop User", "SKU", "Description", "Retail Price", "Quantity"]


@pytest.fixture
def groups(db):
    return {
        name: Group.objects.create(name=name)
        for name in ("managers", "shop_users", "receive_mail")
    }


@pytest.fixture
def manager(groups):
    user = User.objects.create_user(
        username="manager", password="test-pass-123", email="manager@example.com"
    )
    user.groups.add(groups["managers"])
    return user


@pytest.fixture
def shop_user(groups):
    user = User.objects.create_user(
        username="shop1", password="test-pass-123", email="shop1@example.com"
    )
    user.groups.add(groups["shop_users"])
    return user


@pytest.fixture
def other_shop_user(groups):
    user = User.objects.create_user(
        username="shop2", password="test-pass-123", email="shop2@example.com"
    )
    user.groups.add(groups["shop_users"])
    return user


@pytest.fixture
def plain_user(db):
    """An authenticated user belonging to no group."""
    return User.objects.create_user(
        username="nobody", password="test-pass-123", email="nobody@example.com"
    )


@pytest.fixture
def app_config(db):
    """The singleton Admin configuration row most views depend on."""
    return Admin.objects.create(id=1)


@pytest.fixture
def api_client():
    return APIClient()


@pytest.fixture
def manager_client(manager):
    client = APIClient()
    client.force_authenticate(user=manager)
    return client


@pytest.fixture
def shop_client(shop_user):
    client = APIClient()
    client.force_authenticate(user=shop_user)
    return client


@pytest.fixture
def plain_client(plain_user):
    client = APIClient()
    client.force_authenticate(user=plain_user)
    return client


@pytest.fixture
def make_item(db):
    def _make_item(
        sku="SKU-1",
        description="Widget",
        retail_price="9.99",
        quantity=10,
        is_active=True,
    ):
        return Item.objects.create(
            sku=sku,
            description=description,
            retail_price=retail_price,
            quantity=quantity,
            is_active=is_active,
        )

    return _make_item


def build_xlsx(sheets):
    """
    Build an in-memory .xlsx file from {sheet_name: [row, ...]} and return
    a SimpleUploadedFile suitable for posting to the import endpoint.
    """
    workbook = Workbook()
    workbook.remove(workbook.active)
    for name, rows in sheets.items():
        sheet = workbook.create_sheet(title=name)
        for row in rows:
            sheet.append(row)
    output = BytesIO()
    workbook.save(output)
    return SimpleUploadedFile(
        "upload.xlsx",
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def workbook_from_response(response):
    """Load the openpyxl Workbook out of a FileResponse."""
    return load_workbook(BytesIO(b"".join(response.streaming_content)))


def sheet_rows(worksheet):
    """All data rows (skipping the header) of a worksheet as a list of tuples."""
    return list(worksheet.iter_rows(min_row=2, values_only=True))
