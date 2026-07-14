"""
Unit tests for the example custom spreadsheet conversion function
(stock_manager/custom_funcs/spreadsheet_convert.py), which maps a
'Main stock list' sheet into the app's default two-sheet schema.
"""

import pytest
from openpyxl import Workbook

# The custom converter (spreadsheet_convert.py) is user-supplied and gitignored,
# so it is absent from a fresh checkout / deployment. Skip this whole module when
# it isn't present rather than failing collection.
convert_excel = pytest.importorskip(
    "stock_manager.custom_funcs.spreadsheet_convert",
    reason="custom spreadsheet_convert.py not present (user-supplied, gitignored)",
).convert_excel

FULL_HEADER = ["SKUCode", "Item", "BL Price", "Barn", "Feed Barn", "Mengham", "Bog"]


def build_source(rows, header=None):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Main stock list"
    sheet.append(header or FULL_HEADER)
    for row in rows:
        sheet.append(row)
    return workbook


def rows_of(workbook, sheet_name):
    return list(workbook[sheet_name].iter_rows(min_row=2, values_only=True))


class TestConvertExcel:
    def test_produces_both_sheets(self):
        result = convert_excel(
            build_source([["S1", "Desc 1", 10.5, 3, 2, 5, 0]])
        )
        assert set(result.sheetnames) == {"Warehouse Stock", "Shop Stock"}

    def test_warehouse_quantity_sums_locations(self):
        result = convert_excel(
            build_source([["S1", "Desc 1", 10.5, 3, 2, 0, 0]])
        )
        assert rows_of(result, "Warehouse Stock") == [("S1", "Desc 1", 10.5, 5)]

    def test_shop_rows_mapped_to_usernames_and_zero_quantities_dropped(self):
        result = convert_excel(
            build_source([["S1", "Desc 1", 10.5, 0, 0, 5, 0]])
        )
        shop_rows = rows_of(result, "Shop Stock")
        assert shop_rows == [("shop.mengham", "S1", "Desc 1", 10.5, 5)]

    def test_rows_without_sku_dropped(self):
        result = convert_excel(
            build_source(
                [
                    ["S1", "Desc 1", 1, 1, 0, 0, 0],
                    [None, "No sku", 1, 1, 0, 0, 0],
                ]
            )
        )
        warehouse_skus = [row[0] for row in rows_of(result, "Warehouse Stock")]
        assert warehouse_skus == ["S1"]

    def test_missing_price_defaults_to_zero(self):
        result = convert_excel(
            build_source([["S1", "Desc 1", None, 2, 0, 0, 0]])
        )
        assert rows_of(result, "Warehouse Stock") == [("S1", "Desc 1", 0, 2)]

    def test_price_rounded_to_two_decimal_places(self):
        result = convert_excel(
            build_source([["S1", "Desc 1", 10.555, 2, 0, 0, 0]])
        )
        price = rows_of(result, "Warehouse Stock")[0][2]
        assert price == pytest.approx(10.56, abs=0.01)

    def test_duplicate_skus_aggregated_in_warehouse(self):
        result = convert_excel(
            build_source(
                [
                    ["S1", "Desc 1", 1, 2, 0, 0, 0],
                    ["S1", "Desc 1", 1, 3, 0, 0, 0],
                ]
            )
        )
        assert rows_of(result, "Warehouse Stock") == [("S1", "Desc 1", 1, 5)]

    def test_non_numeric_quantity_raises(self):
        with pytest.raises(ValueError, match="NaN"):
            convert_excel(build_source([["S1", "Desc 1", 1, "abc", 0, 0, 0]]))

    def test_warehouse_only_when_no_shop_columns(self):
        header = ["SKUCode", "Item", "BL Price", "Barn"]
        result = convert_excel(
            build_source([["S1", "Desc 1", 1, 2]], header=header)
        )
        assert result.sheetnames == ["Warehouse Stock"]

    def test_shop_only_when_no_warehouse_columns(self):
        header = ["SKUCode", "Item", "BL Price", "Mengham"]
        result = convert_excel(
            build_source([["S1", "Desc 1", 1, 2]], header=header)
        )
        assert result.sheetnames == ["Shop Stock"]
