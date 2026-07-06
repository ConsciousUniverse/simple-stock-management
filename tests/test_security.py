"""
Security regression tests.

Each test asserts the *secure* behaviour for a vulnerability found in the
security audit, so a regression re-introducing the flaw fails the suite.
"""

from io import BytesIO

import pytest
from openpyxl import load_workbook

import ssm.settings as base_settings
from stock_manager.models import Item, ShopItem, TransferItem

pytestmark = pytest.mark.django_db


class TestItemCreateAuthorization:
    """
    Broken access control: POST /api/items/ must be restricted to managers,
    like update and destroy already are. Otherwise any authenticated user
    (e.g. a shop user) can create or silently reactivate+modify warehouse
    stock via a direct API call.
    """

    def test_shop_user_cannot_create_item(self, shop_client):
        response = shop_client.post(
            "/api/items/",
            {"sku": "EVIL", "description": "x", "retail_price": "1.00", "quantity": 1},
            format="json",
        )
        assert response.status_code == 403
        assert not Item.objects.filter(sku="EVIL").exists()

    def test_groupless_user_cannot_create_item(self, plain_client):
        response = plain_client.post(
            "/api/items/",
            {"sku": "EVIL", "description": "x", "retail_price": "1.00", "quantity": 1},
            format="json",
        )
        assert response.status_code == 403
        assert not Item.objects.filter(sku="EVIL").exists()

    def test_shop_user_cannot_reactivate_item_via_create(self, shop_client, make_item):
        make_item(sku="GHOST", description="original", quantity=0, is_active=False)
        response = shop_client.post(
            "/api/items/",
            {"sku": "GHOST", "description": "hijacked", "retail_price": "9.99", "quantity": 999},
            format="json",
        )
        assert response.status_code == 403
        item = Item.objects.get(sku="GHOST")
        assert item.is_active is False
        assert item.description == "original"

    def test_manager_can_still_create_item(self, manager_client):
        response = manager_client.post(
            "/api/items/",
            {"sku": "OK", "description": "Widget", "retail_price": "9.99", "quantity": 5},
            format="json",
        )
        assert response.status_code == 201
        assert Item.objects.filter(sku="OK").exists()


class TestCompleteTransferIDOR:
    """
    IDOR: the cancel path of complete-transfer trusts the shop_user_id in
    the request body. A non-manager must only be able to cancel their *own*
    pending transfers, not another shop user's.
    """

    def payload(self, sku, shop_user_id, cancel="true"):
        return {
            "sku": sku,
            "quantity": 1,
            "shop_user_id": shop_user_id,
            "cancel": cancel,
        }

    def test_shop_user_cannot_cancel_another_users_transfer(
        self, shop_client, shop_user, other_shop_user, app_config, make_item
    ):
        item = make_item(sku="SKU-1", quantity=10)
        victim_transfer = TransferItem.objects.create(
            shop_user=other_shop_user, item=item, quantity=3
        )

        response = shop_client.post(
            "/api/complete-transfer/",
            self.payload("SKU-1", other_shop_user.username),
            format="json",
        )

        assert response.status_code == 403
        assert TransferItem.objects.filter(pk=victim_transfer.pk).exists()

    def test_shop_user_can_still_cancel_own_transfer(
        self, shop_client, shop_user, app_config, make_item
    ):
        item = make_item(sku="SKU-1", quantity=10)
        TransferItem.objects.create(shop_user=shop_user, item=item, quantity=3)

        response = shop_client.post(
            "/api/complete-transfer/",
            self.payload("SKU-1", shop_user.username),
            format="json",
        )

        assert response.status_code == 200
        assert not TransferItem.objects.filter(shop_user=shop_user, item=item).exists()

    def test_manager_can_still_cancel_any_transfer(
        self, manager_client, other_shop_user, app_config, make_item
    ):
        item = make_item(sku="SKU-1", quantity=10)
        TransferItem.objects.create(
            shop_user=other_shop_user, item=item, quantity=3, ordered=True
        )

        response = manager_client.post(
            "/api/complete-transfer/",
            self.payload("SKU-1", other_shop_user.username),
            format="json",
        )

        assert response.status_code == 200
        assert not TransferItem.objects.filter(
            shop_user=other_shop_user, item=item
        ).exists()


class TestExcelFormulaInjection:
    """
    CSV/Excel formula injection: text cells that begin with a formula
    trigger character must be neutralised on export so they are not executed
    when a manager opens the downloaded spreadsheet.
    """

    def workbook_from_response(self, response):
        return load_workbook(BytesIO(b"".join(response.streaming_content)))

    @pytest.mark.parametrize(
        "payload", ["=1+1", "+1+1", "-1+1", "@SUM(A1)", "=cmd|'/c calc'!A0"]
    )
    def test_dangerous_warehouse_cells_are_escaped(
        self, manager_client, make_item, payload
    ):
        make_item(sku="SKU-1", description=payload)

        workbook = self.workbook_from_response(manager_client.get("/api/export_data/"))
        description_cell = workbook["Warehouse Stock"].cell(row=2, column=2).value

        assert description_cell.startswith("'")
        assert description_cell == "'" + payload

    def test_dangerous_shop_cells_are_escaped(
        self, manager_client, shop_user, make_item
    ):
        item = make_item(sku="=evil", description="Widget")
        ShopItem.objects.create(shop_user=shop_user, item=item, quantity=1)

        workbook = self.workbook_from_response(manager_client.get("/api/export_data/"))
        sku_cell = workbook["Shop Stock"].cell(row=2, column=2).value

        assert sku_cell == "'=evil"

    def test_safe_cells_are_untouched(self, manager_client, make_item):
        make_item(sku="SKU-1", description="Ordinary widget")

        workbook = self.workbook_from_response(manager_client.get("/api/export_data/"))
        row = next(
            workbook["Warehouse Stock"].iter_rows(min_row=2, max_row=2, values_only=True)
        )

        assert row[0] == "SKU-1"
        assert row[1] == "Ordinary widget"


class TestReadOnlyStockViewsets:
    """
    Attack-surface reduction: the shop-item and transfer-item collections are
    read-only in the UI (stock is moved via the dedicated transfer/complete
    endpoints), so the viewsets must not expose write verbs.
    """

    def test_shop_items_reject_create(self, shop_client):
        assert shop_client.post("/api/shop_items/", {}, format="json").status_code == 405

    def test_shop_items_reject_delete(self, shop_client, shop_user, make_item):
        ShopItem.objects.create(shop_user=shop_user, item=make_item(sku="SKU-1"), quantity=1)
        assert shop_client.delete("/api/shop_items/SKU-1/").status_code == 405
        assert ShopItem.objects.filter(item__sku="SKU-1").exists()

    def test_transfer_items_reject_create(self, shop_client):
        assert shop_client.post("/api/transfer_items/", {}, format="json").status_code == 405

    def test_transfer_items_reject_delete(self, shop_client, shop_user, make_item):
        TransferItem.objects.create(
            shop_user=shop_user, item=make_item(sku="SKU-1"), quantity=1
        )
        assert shop_client.delete("/api/transfer_items/SKU-1/").status_code == 405
        assert TransferItem.objects.filter(item__sku="SKU-1").exists()

    def test_reads_still_work(self, shop_client):
        assert shop_client.get("/api/shop_items/").status_code == 200
        assert shop_client.get("/api/transfer_items/").status_code == 200


class TestTransportSecuritySettings:
    """
    Cookies and transport must be hardened in real deployments (DEBUG off).
    These assert the production wiring on the base settings module, not the
    test-run overrides.
    """

    def test_secure_cookies_track_debug(self):
        assert base_settings.SESSION_COOKIE_SECURE == (not base_settings.DEBUG)
        assert base_settings.CSRF_COOKIE_SECURE == (not base_settings.DEBUG)

    def test_session_cookie_is_httponly(self):
        assert base_settings.SESSION_COOKIE_HTTPONLY is True

    def test_ssl_redirect_and_hsts_track_debug(self):
        assert base_settings.SECURE_SSL_REDIRECT == (not base_settings.DEBUG)
        if base_settings.DEBUG:
            assert base_settings.SECURE_HSTS_SECONDS == 0
        else:
            assert base_settings.SECURE_HSTS_SECONDS > 0

    def test_proxy_ssl_header_configured(self):
        assert base_settings.SECURE_PROXY_SSL_HEADER == (
            "HTTP_X_FORWARDED_PROTO",
            "https",
        )


class TestContentSecurityPolicy:
    """
    The CSP header must be present and, while it keeps 'unsafe-inline' (the
    templates need inline handlers), it must tightly restrict where scripts
    load from and where the page can connect, and must not permit eval/WASM.
    """

    def get_policy(self, client):
        response = client.get("/accounts/login/")
        assert response.status_code == 200
        assert "Content-Security-Policy" in response
        return response["Content-Security-Policy"]

    def test_header_present_on_responses(self, client):
        assert self.get_policy(client)

    def test_scripts_restricted_to_self_and_known_cdn(self, client):
        policy = self.get_policy(client)
        assert "script-src 'self' 'unsafe-inline' https://code.jquery.com" in policy

    def test_connections_restricted_to_self(self, client):
        # Blocks exfiltration / mining-pool callbacks to external hosts.
        assert "connect-src 'self'" in self.get_policy(client)

    def test_eval_and_wasm_not_allowed(self, client):
        policy = self.get_policy(client)
        assert "unsafe-eval" not in policy
        assert "wasm-unsafe-eval" not in policy

    def test_objects_and_base_locked_down(self, client):
        policy = self.get_policy(client)
        assert "object-src 'none'" in policy
        assert "base-uri 'self'" in policy

    def test_applies_to_authenticated_pages(self, client, shop_user):
        client.force_login(shop_user)
        response = client.get("/")
        assert "Content-Security-Policy" in response


class TestErrorMessagesDoNotLeakInternals:
    """
    Unexpected exceptions must not have their raw text returned to the client
    (CWE-209). Intentional, user-facing validation messages (e.g. "Not enough
    stock") are still shown; only unexpected internal errors are masked.
    """

    SECRET = "internal-detail-hostname-42"

    def test_transfer_item_masks_unexpected_error(
        self, shop_client, app_config, make_item, monkeypatch
    ):
        make_item(quantity=10)

        def boom(*args, **kwargs):
            raise RuntimeError(self.SECRET)

        monkeypatch.setattr("stock_manager.views.transfer_to_shop", boom)
        response = shop_client.post(
            "/api/transfer/",
            {"sku": "SKU-1", "transfer_quantity": "1"},
            format="json",
        )
        assert response.status_code == 400
        assert self.SECRET not in response.json()["detail"]

    def test_transfer_item_still_shows_intended_message(
        self, shop_client, app_config, make_item
    ):
        # A genuine business-rule error must remain visible to the user.
        make_item(quantity=1)
        response = shop_client.post(
            "/api/transfer/",
            {"sku": "SKU-1", "transfer_quantity": "5"},
            format="json",
        )
        assert response.status_code == 400
        assert "Not enough stock" in response.json()["detail"]

    def test_complete_transfer_masks_unexpected_error(
        self, manager_client, shop_user, app_config, make_item, monkeypatch
    ):
        item = make_item(quantity=10)
        TransferItem.objects.create(
            shop_user=shop_user, item=item, quantity=4, ordered=True
        )

        def boom(*args, **kwargs):
            raise RuntimeError(self.SECRET)

        monkeypatch.setattr("stock_manager.views.transfer_to_shop", boom)
        response = manager_client.post(
            "/api/complete-transfer/",
            {"sku": "SKU-1", "quantity": 4, "shop_user_id": "shop1", "cancel": "false"},
            format="json",
        )
        assert response.status_code == 400
        assert self.SECRET not in response.json()["detail"]

    def test_submit_transfer_request_masks_unexpected_error(
        self, shop_client, shop_user, app_config, make_item, monkeypatch
    ):
        TransferItem.objects.create(
            shop_user=shop_user, item=make_item(), quantity=2
        )

        def boom(*args, **kwargs):
            raise RuntimeError(self.SECRET)

        # SendEmail.compose is invoked inside the try; make it blow up.
        monkeypatch.setattr(
            "stock_manager.views.SendEmail.compose", boom
        )
        response = shop_client.post(
            "/api/submit-transfer-request/", {}, format="json"
        )
        assert response.status_code == 400
        assert self.SECRET not in response.json()["detail"]
